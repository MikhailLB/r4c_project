import io
from openpyxl import Workbook


def generate_excel(weekly_summary):

    """Создает Excel-файл со сводкой производства"""
# Тут распишу подробнее, потому что, как по мне, это сложная логика

    wb = Workbook()

    # Достаем все УНИКАЛЬНЫЕ названия моделей
    models = set(entry['model'] for entry in weekly_summary)

    if not models:
        # Если нет данных, создаем лист с No data

        ws = wb.active
        ws.title = "No Data"
        ws.append(["Модель", "Версия", "Количество за неделю"])
    else:
        # Если данные есть, генерируем листы для каждой модели
        for model in models:
            ws = wb.create_sheet(title=f"Model {model}")
            ws.append(["Модель", "Версия", "Количество за неделю"])

            # Пробегаемся по queryset, если находим там название модели равное текущему в цикле, добавляем на страничку данные о модели
            for entry in weekly_summary:
                if entry['model'] == model:
                    ws.append([entry['model'], entry['version'], entry['count']])

        # Удаляем стандартный лист, если он существует
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    # Сохраняем в буфер, а не на локальной машине, чтобы отправить результат через HTTP
    buffer = io.BytesIO()
    wb.save(buffer)

    # Ставим курсор на начало, чтобы можно было вернуть buffer
    buffer.seek(0)
    return buffer